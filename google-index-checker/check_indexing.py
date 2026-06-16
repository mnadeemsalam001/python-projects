import time
from datetime import datetime
from pathlib import Path

from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_DIR         = Path(__file__).parent   # always the folder this script lives in
CREDENTIALS_FILE = BASE_DIR / "credentials.json"
EXCEL_FILE       = BASE_DIR / "Index_Tracker.xlsx"
SITE_URL         = "https://datamaven360.com/"
SCOPES           = ["https://www.googleapis.com/auth/webmasters.readonly"]
DELAY_SECONDS    = 1   # pause between API calls (rate limit: 600 req/min)
# ──────────────────────────────────────────────────────────────────────────────

# Cell fill colors
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_DATE  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def build_service():
    credentials = service_account.Credentials.from_service_account_file(
        CREDENTIALS_FILE, scopes=SCOPES
    )
    return build("searchconsole", "v1", credentials=credentials)


def check_url(service, url: str) -> dict:
    try:
        response = service.urlInspection().index().inspect(
            body={"inspectionUrl": url, "siteUrl": SITE_URL}
        ).execute()
        result = response.get("inspectionResult", {}).get("indexStatusResult", {})
        return {
            "verdict":       result.get("verdict", "UNKNOWN"),
            "coverageState": result.get("coverageState", "Unknown"),
        }
    except Exception as e:
        return {"verdict": "ERROR", "coverageState": f"Error: {str(e)}"}


def load_urls(ws) -> list:
    return [
        row[1].strip()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True)
        if row[1]
    ]


def get_or_create_date_column(ws, today: str) -> int:
    """Find today's column if it exists, otherwise add a new one."""
    for cell in ws[1]:
        if cell.value == today:
            return cell.column
    next_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=next_col, value=today)
    header_cell.fill = FILL_DATE
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center")
    return next_col


def get_fill(verdict: str) -> PatternFill:
    if verdict == "PASS":
        return FILL_GREEN
    elif verdict == "FAIL":
        return FILL_RED
    return FILL_GRAY


def main():
    today = datetime.now().strftime("%Y-%m-%d")

    print()
    print("=" * 68)
    print("  GOOGLE INDEX CHECKER — DataMaven360")
    print(f"  Date : {today}")
    print("=" * 68)

    # Validate files exist
    if not Path(EXCEL_FILE).exists():
        print(f"\nERROR: '{EXCEL_FILE}' not found.")
        print("Place the Excel file in the same folder as this script.\n")
        return

    if not Path(CREDENTIALS_FILE).exists():
        print(f"\nERROR: '{CREDENTIALS_FILE}' not found.")
        print("Download it from Google Cloud Console → Service Account → Keys.\n")
        return

    # Load workbook and URLs
    wb   = openpyxl.load_workbook(EXCEL_FILE)
    ws   = wb.active
    urls = load_urls(ws)

    if not urls:
        print("\nNo URLs found. Make sure Column A has URLs starting from row 2.\n")
        return

    print(f"\nFound {len(urls)} URL(s). Connecting to Search Console API...")

    try:
        service = build_service()
        print("Connected.\n")
    except Exception as e:
        print(f"\nERROR: Could not connect to API.\n{e}\n")
        return

    print("-" * 68)

    col_index = get_or_create_date_column(ws, today)
    counts = {"Indexed": 0, "Not Indexed": 0, "Error": 0}

    for i, url in enumerate(urls, 1):
        print(f"[{i:02d}/{len(urls):02d}] {url}")

        result   = check_url(service, url)
        verdict  = result["verdict"]
        coverage = result["coverageState"]

        if verdict == "PASS":
            counts["Indexed"] += 1
        elif verdict == "ERROR":
            counts["Error"] += 1
        else:
            counts["Not Indexed"] += 1

        cell = ws.cell(row=i + 1, column=col_index, value=coverage)
        cell.fill      = get_fill(verdict)
        cell.alignment = Alignment(horizontal="center")

        print(f"        {coverage}\n")

        if i < len(urls):
            time.sleep(DELAY_SECONDS)

    wb.save(EXCEL_FILE)

    print("-" * 68)
    print("SUMMARY")
    print(f"  Indexed     : {counts['Indexed']}")
    print(f"  Not Indexed : {counts['Not Indexed']}")
    print(f"  Errors      : {counts['Error']}")
    print(f"  Total       : {len(urls)}")
    print("-" * 68)
    print(f"\nResults saved to: {EXCEL_FILE}\n")


if __name__ == "__main__":
    main()
