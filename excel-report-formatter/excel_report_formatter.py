"""
================================================================================
Excel Report Formatter — Standalone Demo
================================================================================
Author : Muhammad Nadeem Salam
Version: 1.0

Description:
    Demonstrates professional Excel report formatting using openpyxl.
    Contains an embedded sample dataset — no external files required.

    All formatting utilities are extracted from a real-world reporting project
    and are ready to adapt: swap in your own DataFrame and adjust the column
    definitions to match your data.

Sheets produced:
    1. Sales Report      — Full report with group headers, performance-tier
                           badges, alternating rows, conditional formatting,
                           and data bars
    2. Regional Summary  — Aggregated rollup showing merged group headers,
                           color-coded variance, and grand totals

Requirements:
    pip install pandas openpyxl

Output:
    sample_report.xlsx  (saved in the same folder as this script)

================================================================================
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


# ================================================================================
# COLOR PALETTE
# ================================================================================
# All colors as 6-digit hex strings (no leading #).
# Organized in light/dark pairs so badges and backgrounds stay consistent.

C = dict(
    dark      = "0D1B2A",   # near-black navy  — title bars, primary text
    teal      = "0D9488",   # primary accent    — volume group headers
    teal_lt   = "E1F5EE",   # teal bg tint      — volume data cells
    teal_dk   = "085041",   # dark teal text    — Top Performer badge text
    purple    = "7C3AED",   # performance group headers
    purple_lt = "EEEDFE",   # performance data cells
    gold      = "F59E0B",
    gold_lt   = "FAEEDA",
    gold_dk   = "633806",
    coral     = "F43F5E",   # negative / warning
    coral_lt  = "FFE4E6",
    coral_dk  = "9F1239",
    green     = "16A34A",
    green_lt  = "DCFCE7",
    green_dk  = "14532D",
    amber     = "B45309",   # status group headers
    amber_lt  = "FEF3C7",
    slate     = "334155",   # default body text
    white     = "FFFFFF",
    off_w     = "F8FAFC",   # alternating row — even rows
)


# ================================================================================
# STYLE HELPERS
# ================================================================================
# Small, composable functions that return openpyxl style objects.
# Use them individually or combine via set_cell() for one-liner formatting.

def fill(hex_color):
    """Solid cell fill from a 6-digit hex string."""
    return PatternFill("solid", fgColor=hex_color)

def fnt(bold=False, color="000000", size=10, italic=False, name="Arial"):
    """Build a Font object."""
    return Font(bold=bold, color=color, name=name, size=size, italic=italic)

def bdr(color="E2E8F0", style="thin"):
    """Four-sided border with a uniform color and style."""
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def aln(h="left", v="center", wrap=False):
    """Build an Alignment object."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(cell, value, fmt="General", bold=False, color=None,
             bg=None, h_align="center", v_align="center",
             wrap=False, border_color="E2E8F0", italic=False):
    """
    One-call cell formatter.
    Sets value, font, fill, alignment, border, and number format in a single call.
    """
    cell.value         = value
    cell.font          = fnt(bold=bold, color=color or C["slate"], italic=italic)
    cell.fill          = fill(bg or C["off_w"])
    cell.alignment     = aln(h=h_align, v=v_align, wrap=wrap)
    cell.border        = bdr(color=border_color)
    cell.number_format = fmt

def group_header(ws, row, col_start, col_end, label, bg, fg):
    """
    Merged group-header cell spanning col_start to col_end.
    Draws a single wide label above a column group (e.g. "Sales Volume").
    """
    start_ref = get_column_letter(col_start) + str(row)
    end_ref   = get_column_letter(col_end)   + str(row)
    if col_start != col_end:
        ws.merge_cells(f"{start_ref}:{end_ref}")
    cell           = ws[start_ref]
    cell.value     = label
    cell.font      = fnt(bold=True, color=fg, size=10)
    cell.fill      = fill(bg)
    cell.alignment = aln("center")
    cell.border    = bdr(color=bg)   # border matches bg so merged look is seamless

def col_header(ws, row, col_idx, text, bg, fg, width=None):
    """Single column header cell with optional column-width setting."""
    cell           = ws.cell(row, col_idx, text)
    cell.font      = fnt(bold=True, color=fg, size=9)
    cell.fill      = fill(bg)
    cell.alignment = aln("center", wrap=True)
    cell.border    = bdr(color=C["white"])
    if width:
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def title_bar(ws, row, text, span, height=36, size=14):
    """Full-width dark title bar with white bold text."""
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    cell           = ws[f"A{row}"]
    cell.value     = "  " + text          # leading spaces for left padding
    cell.font      = fnt(bold=True, color=C["white"], size=size)
    cell.fill      = fill(C["dark"])
    cell.alignment = aln("left")

def subtitle_bar(ws, row, text, span, height=18):
    """Light-gray subtitle / note row with small italic text."""
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    cell           = ws[f"A{row}"]
    cell.value     = "  " + text
    cell.font      = fnt(italic=True, color="64748B", size=9)
    cell.fill      = fill("F8FAFC")
    cell.alignment = aln("left")

def totals_row(ws, row, n_cols, fill_color=None):
    """Apply dark-background + teal-border styling to every cell in the totals row."""
    for ci in range(1, n_cols + 1):
        cell           = ws.cell(row, ci)
        cell.font      = fnt(bold=True, color=C["white"], size=10)
        cell.fill      = fill(fill_color or C["dark"])
        cell.alignment = aln("center")
        cell.border    = bdr(color=C["teal"], style="medium")

def row_bg(row_index):
    """Alternating row background: off-white on even rows, white on odd rows."""
    return C["off_w"] if row_index % 2 == 0 else C["white"]


# ================================================================================
# PERFORMANCE TIER CONFIG
# ================================================================================

TIER_COLORS = {
    "Top Performer":   (C["teal_lt"],  C["teal_dk"]),
    "Above Average":   (C["green_lt"], C["green_dk"]),
    "Below Average":   (C["gold_lt"],  C["gold_dk"]),
    "Needs Attention": (C["coral_lt"], C["coral_dk"]),
}

def perf_tier(pct_of_target):
    """Assign a performance tier label based on % of sales target achieved."""
    if pct_of_target >= 1.10: return "Top Performer"
    if pct_of_target >= 0.95: return "Above Average"
    if pct_of_target >= 0.80: return "Below Average"
    return "Needs Attention"


# ================================================================================
# SAMPLE DATASET  (replace this block with your own pd.read_* call)
# ================================================================================

raw = {
    "ID":        [1,        2,          3,          4,          5,
                  6,        7,          8,          9,          10],
    "Region":    ["North",  "North",    "South",    "South",    "East",
                  "East",   "West",     "West",     "Central",  "Central"],
    "Sales Rep": ["Alice Morgan",   "Brian Osei",   "Carla Diaz",   "David Kim",
                  "Elena Frost",    "Felix Ramos",  "Grace Liu",    "Hank Patel",
                  "Iris Chen",      "James Okafor"],
    "Q1 Target": [120000, 95000, 140000, 105000, 130000,
                   88000, 115000, 100000,  97000,  75000],
    "Q1 Sales":  [134500,  88200, 152300, 101800, 128900,
                   96400, 107600, 118500,  80200,  84700],
    "Units Sold":[1345,    882,   1523,   1018,   1289,
                   964,   1076,   1185,    802,    847],
    "Growth %":  [ 0.08, -0.04,   0.15,   0.03,  -0.02,
                   0.21,  0.06,   0.12,  -0.07,   0.18],
}

df = pd.DataFrame(raw)
df["% of Target"]   = (df["Q1 Sales"] / df["Q1 Target"]).round(4)
df["Avg Deal Size"] = (df["Q1 Sales"] / df["Units Sold"]).round(2)
df["Perf Tier"]     = df["% of Target"].apply(perf_tier)
df["Variance"]      = df["Q1 Sales"] - df["Q1 Target"]

# Regional rollup for Sheet 2
summary = (
    df.groupby("Region", as_index=False)
    .agg(Reps=("Sales Rep", "count"), Target=("Q1 Target", "sum"),
         Sales=("Q1 Sales", "sum"),   Units=("Units Sold", "sum"))
)
summary["% of Target"] = (summary["Sales"] / summary["Target"]).round(4)
summary["Variance"]    = summary["Sales"] - summary["Target"]

n  = len(df)
ns = len(summary)

print(f"Dataset loaded: {n} reps across {summary['Region'].nunique()} regions")


# ================================================================================
# BUILD WORKBOOK
# ================================================================================

print("Building workbook...")
wb = Workbook()
wb.remove(wb.active)   # delete the default empty sheet


# ============================================================================
# SHEET 1: SALES REPORT  — full rep-level detail
# ============================================================================

ws1 = wb.create_sheet("Sales Report")
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = C["teal"]

SPAN = 11   # total columns on this sheet

# ── Title & subtitle ─────────────────────────────────────────────────────────
title_bar(ws1, 1, "Q1 Sales Performance Report  |  Sample Dataset", SPAN)
subtitle_bar(ws1, 2,
    "Perf Tier: ≥110% = Top Performer · ≥95% = Above Average · "
    "≥80% = Below Average · <80% = Needs Attention", SPAN)
ws1.row_dimensions[3].height = 6   # thin spacer row

# ── Group headers (row 4) ─────────────────────────────────────────────────────
ws1.row_dimensions[4].height = 22
for label, cs, ce, bg, fg in [
    ("Identity",     1,  3,  C["dark"],   C["white"]),
    ("Sales Volume", 4,  6,  C["teal"],   C["white"]),
    ("Performance",  7,  9,  C["purple"], C["white"]),
    ("Status",      10, 11,  C["amber"],  C["white"]),
]:
    group_header(ws1, 4, cs, ce, label, bg, fg)

# ── Column headers (row 5) ───────────────────────────────────────────────────
ws1.row_dimensions[5].height = 46
for ci, (label, width, bg, fg) in enumerate([
    ("ID",              5,  C["dark"],   C["white"]),
    ("Sales Rep",      22,  C["dark"],   C["white"]),
    ("Region",         10,  C["dark"],   C["white"]),
    ("Q1 Target",      13,  C["teal"],   C["white"]),
    ("Q1 Sales",       13,  C["teal"],   C["white"]),
    ("Units\nSold",    10,  C["teal"],   C["white"]),
    ("% of\nTarget",   11,  C["purple"], C["white"]),
    ("Avg Deal\nSize", 13,  C["purple"], C["white"]),
    ("Growth %",       10,  C["purple"], C["white"]),
    ("Variance",       13,  C["amber"],  C["dark"]),
    ("Perf Tier",      15,  C["amber"],  C["dark"]),
], start=1):
    col_header(ws1, 5, ci, label, bg, fg, width=width)

# ── Data rows ────────────────────────────────────────────────────────────────
for ri, row in df.iterrows():
    er  = ri + 6
    rb  = row_bg(ri)
    ws1.row_dimensions[er].height = 18

    tier_bg, tier_fg = TIER_COLORS[row["Perf Tier"]]
    positive         = row["Variance"] >= 0

    def w(ci, val, fmt="General", bg=None, col=None, bold=False):
        c = ws1.cell(er, ci, val)
        c.font      = fnt(bold=bold, color=col or (C["dark"] if ci <= 3 else C["slate"]))
        c.fill      = fill(bg or rb)
        c.alignment = aln("left" if ci == 2 else "center")
        c.border    = bdr()
        c.number_format = fmt

    w(1, int(row["ID"]),         col="94A3B8")
    w(2, row["Sales Rep"],       bold=True)
    w(3, row["Region"])
    w(4, row["Q1 Target"],       "$#,##0",    bg=C["teal_lt"])
    w(5, row["Q1 Sales"],        "$#,##0",    bg=C["teal_lt"])
    w(6, int(row["Units Sold"]), "#,##0",     bg=C["teal_lt"])
    w(7, row["% of Target"],     "0.0%",      bg=C["purple_lt"])
    w(8, row["Avg Deal Size"],   "$#,##0.00", bg=C["purple_lt"])

    # Growth % — green if positive, coral if negative
    g = row["Growth %"]
    c9            = ws1.cell(er, 9, g)
    c9.font       = fnt(bold=True, color=C["green_dk"] if g >= 0 else C["coral_dk"])
    c9.fill       = fill(C["green_lt"] if g >= 0 else C["coral_lt"])
    c9.alignment  = aln("center")
    c9.border     = bdr()
    c9.number_format = "+0.0%;-0.0%;0.0%"

    # Variance — green if hit target, coral if missed
    c10            = ws1.cell(er, 10, row["Variance"])
    c10.font       = fnt(bold=True, color=C["green_dk"] if positive else C["coral_dk"])
    c10.fill       = fill(C["green_lt"] if positive else C["coral_lt"])
    c10.alignment  = aln("center")
    c10.border     = bdr()
    c10.number_format = "+$#,##0;-$#,##0;$0"

    # Performance tier badge
    c11            = ws1.cell(er, 11, row["Perf Tier"])
    c11.font       = fnt(bold=True, color=tier_fg, size=9)
    c11.fill       = fill(tier_bg)
    c11.alignment  = aln("center")
    c11.border     = bdr()

# ── Totals row ───────────────────────────────────────────────────────────────
tr = 6 + n
totals_row(ws1, tr, SPAN)
ws1.cell(tr, 1).value = "TOTALS"
ws1.merge_cells(f"A{tr}:C{tr}")

for ci, (col_letter, fmt) in enumerate([
    ("D", "$#,##0"),   # Q1 Target
    ("E", "$#,##0"),   # Q1 Sales
    ("F", "#,##0"),    # Units Sold
], start=4):
    c = ws1.cell(tr, ci, f"=SUM({col_letter}6:{col_letter}{tr-1})")
    c.font, c.fill         = fnt(bold=True, color=C["white"]), fill(C["dark"])
    c.alignment, c.border  = aln("center"), bdr(C["teal"], "medium")
    c.number_format        = fmt

c_avg = ws1.cell(tr, 7, f"=AVERAGE(G6:G{tr-1})")
c_avg.font, c_avg.fill       = fnt(bold=True, color=C["white"]), fill(C["dark"])
c_avg.alignment, c_avg.border = aln("center"), bdr(C["teal"], "medium")
c_avg.number_format           = "0.0%"

c_var = ws1.cell(tr, 10, f"=SUM(J6:J{tr-1})")
c_var.font, c_var.fill       = fnt(bold=True, color=C["white"]), fill(C["dark"])
c_var.alignment, c_var.border = aln("center"), bdr(C["teal"], "medium")
c_var.number_format           = "+$#,##0;-$#,##0;$0"

# ── Conditional formatting ───────────────────────────────────────────────────
# Color scale: red → yellow → green on % of Target column
ws1.conditional_formatting.add(
    f"G6:G{tr-1}",
    ColorScaleRule(
        start_type="min",       start_color="FFE4E6",
        mid_type="percentile",  mid_value=50, mid_color="FEF3C7",
        end_type="max",         end_color="DCFCE7",
    ),
)
# Data bar on Q1 Sales column
ws1.conditional_formatting.add(
    f"E6:E{tr-1}",
    DataBarRule(start_type="min", start_value=0,
                end_type="max",   end_value=None, color=C["teal"]),
)

ws1.freeze_panes = "D6"   # freeze identity columns and header rows


# ============================================================================
# SHEET 2: REGIONAL SUMMARY  — aggregated rollup
# ============================================================================

ws2 = wb.create_sheet("Regional Summary")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = C["purple"]

SPAN2 = 7

# ── Title & subtitle ─────────────────────────────────────────────────────────
title_bar(ws2, 1, "Regional Summary  |  Q1 Sales Performance", SPAN2)
subtitle_bar(ws2, 2, "Aggregated from individual rep data on the Sales Report sheet.", SPAN2)
ws2.row_dimensions[3].height = 8

# ── Group headers ────────────────────────────────────────────────────────────
ws2.row_dimensions[4].height = 22
for label, cs, ce, bg, fg in [
    ("Region",       1, 1, C["dark"],   C["white"]),
    ("Sales Volume", 2, 5, C["teal"],   C["white"]),
    ("Performance",  6, 7, C["purple"], C["white"]),
]:
    group_header(ws2, 4, cs, ce, label, bg, fg)

# ── Column headers ───────────────────────────────────────────────────────────
ws2.row_dimensions[5].height = 40
for ci, (label, width, bg, fg) in enumerate([
    ("Region",        14, C["dark"],   C["white"]),
    ("Reps",           7, C["teal"],   C["white"]),
    ("Q1 Target",     14, C["teal"],   C["white"]),
    ("Q1 Sales",      14, C["teal"],   C["white"]),
    ("Units",         10, C["teal"],   C["white"]),
    ("% of\nTarget",  12, C["purple"], C["white"]),
    ("Variance",      14, C["purple"], C["white"]),
], start=1):
    col_header(ws2, 5, ci, label, bg, fg, width=width)

# ── Data rows ────────────────────────────────────────────────────────────────
for ri, row in summary.iterrows():
    er  = ri + 6
    rb  = row_bg(ri)
    ws2.row_dimensions[er].height = 18
    pct = row["% of Target"]

    def w2(ci, val, fmt="General", bg=None, bold=False, col=None):
        c = ws2.cell(er, ci, val)
        c.font      = fnt(bold=bold, color=col or (C["dark"] if ci == 1 else C["slate"]))
        c.fill      = fill(bg or rb)
        c.alignment = aln("left" if ci == 1 else "center")
        c.border    = bdr()
        c.number_format = fmt

    w2(1, row["Region"],      bold=True)
    w2(2, int(row["Reps"]),   "#,##0",  bg=C["teal_lt"])
    w2(3, row["Target"],      "$#,##0", bg=C["teal_lt"])
    w2(4, row["Sales"],       "$#,##0", bg=C["teal_lt"])
    w2(5, int(row["Units"]),  "#,##0",  bg=C["teal_lt"])

    # % of Target badge — colored by tier
    t_bg, t_fg = TIER_COLORS[perf_tier(pct)]
    c6            = ws2.cell(er, 6, pct)
    c6.font       = fnt(bold=True, color=t_fg)
    c6.fill       = fill(t_bg)
    c6.alignment  = aln("center")
    c6.border     = bdr()
    c6.number_format = "0.0%"

    # Variance
    var = row["Variance"]
    c7           = ws2.cell(er, 7, var)
    c7.font      = fnt(bold=True, color=C["green_dk"] if var >= 0 else C["coral_dk"])
    c7.fill      = fill(C["green_lt"] if var >= 0 else C["coral_lt"])
    c7.alignment = aln("center")
    c7.border    = bdr()
    c7.number_format = "+$#,##0;-$#,##0;$0"

# ── Totals row ───────────────────────────────────────────────────────────────
tr2 = 6 + ns
totals_row(ws2, tr2, SPAN2)
ws2.cell(tr2, 1).value = "GRAND TOTAL"

for ci, (col_letter, fmt) in enumerate([
    ("C", "$#,##0"),   # Q1 Target
    ("D", "$#,##0"),   # Q1 Sales
    ("E", "#,##0"),    # Units
], start=3):
    c = ws2.cell(tr2, ci, f"=SUM({col_letter}6:{col_letter}{tr2-1})")
    c.font, c.fill        = fnt(bold=True, color=C["white"]), fill(C["dark"])
    c.alignment, c.border = aln("center"), bdr(C["teal"], "medium")
    c.number_format       = fmt

c6t = ws2.cell(tr2, 6, f"=AVERAGE(F6:F{tr2-1})")
c6t.font, c6t.fill        = fnt(bold=True, color=C["white"]), fill(C["dark"])
c6t.alignment, c6t.border = aln("center"), bdr(C["teal"], "medium")
c6t.number_format         = "0.0%"

c7t = ws2.cell(tr2, 7, f"=SUM(G6:G{tr2-1})")
c7t.font, c7t.fill        = fnt(bold=True, color=C["white"]), fill(C["dark"])
c7t.alignment, c7t.border = aln("center"), bdr(C["teal"], "medium")
c7t.number_format         = "+$#,##0;-$#,##0;$0"

# Color scale on % of Target
ws2.conditional_formatting.add(
    f"F6:F{tr2-1}",
    ColorScaleRule(
        start_type="min",       start_color="FFE4E6",
        mid_type="percentile",  mid_value=50, mid_color="FEF3C7",
        end_type="max",         end_color="DCFCE7",
    ),
)

ws2.freeze_panes = "B6"


# ================================================================================
# SAVE
# ================================================================================

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_report.xlsx")
wb.save(OUTPUT)
print(f"Report saved: {OUTPUT}")
